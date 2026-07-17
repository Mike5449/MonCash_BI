"""
Streaming export helpers (XLSX + CSV).

Permet d'exposer un endpoint qui génère un fichier Excel côté serveur
(via openpyxl en mode write-only — mémoire bornée) et le renvoie
directement au navigateur en StreamingResponse.

Avantages vs export JSON + XLSX côté front :
- Pas de double parsing (JSON → JS objects → XLSX en mémoire navigateur)
- Pas de freeze du navigateur sur gros volumes (1M+ lignes)
- Header X-Row-Count exposé → progression utilisable côté front
"""

import io
import csv
from typing import Any, Callable, List, Tuple, Iterable
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment


# Colonne = (key, label, picker)
# - key : identifiant interne (utilisé pour filtrer côté front)
# - label : en-tête affiché dans Excel
# - picker : fonction qui extrait la valeur depuis un dict row
ExcelColumn = Tuple[str, str, Callable[[dict], Any]]

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MEDIA_TYPE  = "text/csv; charset=utf-8"


def select_columns(all_columns: List[ExcelColumn], requested_keys: str | None) -> List[ExcelColumn]:
    """
    Filtre `all_columns` selon la liste comma-separated `requested_keys`.
    Vide ou None → renvoie toutes les colonnes.
    """
    keys = set([c.strip() for c in (requested_keys or "").split(",") if c.strip()])
    if not keys:
        return all_columns
    filtered = [c for c in all_columns if c[0] in keys]
    return filtered if filtered else all_columns


def stream_xlsx(
    rows: List[dict],
    columns: List[ExcelColumn],
    filename: str,
    sheet_name: str = "Sheet1",
) -> StreamingResponse:
    """
    Construit un Workbook openpyxl write-only et le streame.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name[:31])  # Excel limite le nom de sheet à 31 chars

    # En-tête stylé (fond bleu nuit, texte blanc, gras)
    header_cells = []
    for _, label, _ in columns:
        cell = WriteOnlyCell(ws, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cells.append(cell)
    ws.append(header_cells)

    # Largeurs approximatives (basées sur le label de l'en-tête)
    for i, (_, label, _) in enumerate(columns, start=1):
        if i <= 26:
            col_letter = chr(ord('A') + (i - 1))
        else:
            # AA, AB, … (jusqu'à ZZ)
            col_letter = chr(ord('A') + ((i - 1) // 26 - 1)) + chr(ord('A') + ((i - 1) % 26))
        ws.column_dimensions[col_letter].width = max(12, len(label) + 2)

    # Données
    for r in rows:
        ws.append([pick(r) for _, _, pick in columns])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
            # Permet au client de lire ces headers même en CORS
            "Access-Control-Expose-Headers": "Content-Disposition, X-Row-Count, Content-Length",
        },
    )


def stream_csv(
    rows: Iterable[dict],
    columns: List[ExcelColumn],
    filename: str,
    batch_size: int = 5000,
) -> StreamingResponse:
    """
    Streame un CSV par BATCHES de `batch_size` lignes (vs ligne-par-ligne) pour réduire
    drastiquement l'overhead Python : 800k yields = ~800k thread context switches via
    anyio (très lent). En batchant à 5000, on a ~160 yields → 5000× plus rapide.

    BOM UTF-8 inclus pour qu'Excel interprète correctement les caractères accentués.
    """
    labels = [label for _, label, _ in columns]
    pickers = [pick for _, _, pick in columns]

    def generate():
        # BOM UTF-8 → Excel lit correctement les caractères accentués
        yield "﻿".encode("utf-8")

        # Header — yield immédiat (le navigateur reçoit déjà des bytes)
        head_buf = io.StringIO()
        csv.writer(head_buf, lineterminator="\n").writerow(labels)
        yield head_buf.getvalue().encode("utf-8")

        # Body — batché pour réduire l'overhead de yield/encode
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        n_pickers = len(pickers)
        count = 0

        for r in rows:
            # Inline picker loop (plus rapide qu'une list comprehension pour petits N)
            row_vals = [None] * n_pickers
            for i in range(n_pickers):
                row_vals[i] = pickers[i](r)
            writer.writerow(row_vals)
            count += 1
            if count >= batch_size:
                yield buf.getvalue().encode("utf-8")
                buf.seek(0); buf.truncate(0)
                count = 0

        # Flush du dernier batch partiel
        if count > 0:
            yield buf.getvalue().encode("utf-8")

    return StreamingResponse(
        generate(),
        media_type=CSV_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
